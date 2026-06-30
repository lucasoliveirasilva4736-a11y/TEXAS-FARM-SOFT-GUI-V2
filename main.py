import traceback
import tkinter as tk
import tkinter.messagebox as tkmb
import os
import re
import math

# Bloco Try-Except global para capturar erros caso falte alguma biblioteca
try:
    import customtkinter as ctk
    from tkinter import messagebox, ttk, filedialog
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

    # Configuração visual do sistema
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    # =======================================================
    # MÓDULO EXTRA: GERENCIADOR DE CORES
    # =======================================================
    class JanelaGerenciadorCores(ctk.CTkToplevel):
        def __init__(self, master):
            super().__init__(master)
            self.title("TEXAS FARM - Base de Cores do Sistema")
            self.geometry("1200x750")
            self.attributes("-topmost", True)
            
            # -------------------------------------------------------------
            # !!! AJUSTE AQUI O CAMINHO EXATO DO SEU ARQUIVO DE CORES !!!
            # -------------------------------------------------------------
            self.caminho_arquivo_cores = r"Z:\Programação Corte\Cores_Modelos_PCP_Atualizado.xlsx" 
            # Se for dentro de uma pasta, mude para: r"Z:\Programação Corte\NOME_DO_ARQUIVO.xlsx"
            
            self.dados_cores = {}
            
            # --- CONSTRUÇÃO DA INTERFACE ---
            self.frame_topo = ctk.CTkFrame(self)
            self.frame_topo.pack(pady=10, padx=20, fill="x")
            
            ctk.CTkLabel(self.frame_topo, text="🎨 BANCO DE DADOS DE CORES E MODELOS", font=("Roboto", 18, "bold"), text_color="#2fa572").pack(pady=10)
            
            self.lbl_arquivo = ctk.CTkLabel(self.frame_topo, text=f"Arquivo Vinculado: {self.caminho_arquivo_cores}", text_color="gray", font=("Roboto", 12))
            self.lbl_arquivo.pack()
            
            # Layout Principal (Tabela à esquerda, Formulário à direita)
            self.frame_corpo = ctk.CTkFrame(self, fg_color="transparent")
            self.frame_corpo.pack(fill="both", expand=True, padx=20, pady=5)
            
            # ESQUERDA: Tabela de Cores
            self.frame_esq = ctk.CTkFrame(self.frame_corpo)
            self.frame_esq.pack(side="left", fill="both", expand=True, padx=(0, 10))
            
            self.frame_filtro = ctk.CTkFrame(self.frame_esq, fg_color="transparent")
            self.frame_filtro.pack(fill="x", pady=10, padx=10)
            
            ctk.CTkLabel(self.frame_filtro, text="Visualizar Modelo (Aba):", font=("Roboto", 13, "bold")).pack(side="left")
            self.menu_abas = ctk.CTkOptionMenu(self.frame_filtro, values=["-"], command=self.exibir_cores_aba)
            self.menu_abas.pack(side="left", padx=10)
            
            style = ttk.Style()
            style.theme_use("clam")
            style.configure("Treeview.Cores", font=("Roboto", 12), rowheight=30, background="#2b2b2b", foreground="white")
            style.configure("Treeview.Heading", font=("Roboto", 13, "bold"), background="#8B4513", foreground="white")
            
            self.colunas = ("Modelo", "Tecido", "Cor (Comercial)", "Nome no Sistema (Parênteses)")
            self.tabela = ttk.Treeview(self.frame_esq, columns=self.colunas, show="headings", style="Treeview.Cores")
            
            for col in self.colunas:
                self.tabela.heading(col, text=col)
                largura = 120 if col == "Modelo" else (180 if col == "Tecido" else 220)
                self.tabela.column(col, width=largura, anchor="center")
            self.tabela.pack(fill="both", expand=True, padx=10, pady=(0,10))
            
            # DIREITA: Painel Duplo (Criar Modelo / Criar Cor)
            self.frame_dir = ctk.CTkScrollableFrame(self.frame_corpo, width=400)
            self.frame_dir.pack(side="right", fill="y")
            
            # --- SEÇÃO 1: CRIAR NOVO MODELO (ABA) ---
            self.frame_novo_mod = ctk.CTkFrame(self.frame_dir)
            self.frame_novo_mod.pack(fill="x", pady=(0, 20), padx=5)
            
            ctk.CTkLabel(self.frame_novo_mod, text="✨ 1. CRIAR NOVO MODELO (ABA)", font=("Roboto", 15, "bold"), text_color="#3b8ed0").pack(pady=(15,10))
            
            ctk.CTkLabel(self.frame_novo_mod, text="Nome do Novo Modelo:", font=("Roboto", 12)).pack(anchor="w", padx=20)
            self.entry_novo_modelo_aba = ctk.CTkEntry(self.frame_novo_mod, placeholder_text="Ex: Jaqueta")
            self.entry_novo_modelo_aba.pack(fill="x", padx=20, pady=(0, 10))
            
            self.btn_criar_aba = ctk.CTkButton(self.frame_novo_mod, text="✚ CRIAR ABA NO EXCEL", height=40, font=("Roboto", 13, "bold"), fg_color="#1D6F42", hover_color="#144d2e", command=self.criar_nova_aba)
            self.btn_criar_aba.pack(fill="x", padx=20, pady=(0, 20))
            
            # --- SEÇÃO 2: CADASTRAR NOVA COR ---
            self.frame_nova_cor = ctk.CTkFrame(self.frame_dir)
            self.frame_nova_cor.pack(fill="x", padx=5)
            
            ctk.CTkLabel(self.frame_nova_cor, text="🎨 2. CADASTRAR NOVA COR", font=("Roboto", 15, "bold"), text_color="#d08d3b").pack(pady=(15,10))
            
            ctk.CTkLabel(self.frame_nova_cor, text="Escolher Modelo (Aba Existente):", font=("Roboto", 12)).pack(anchor="w", padx=20)
            self.combo_modelo_novo = ctk.CTkOptionMenu(self.frame_nova_cor, values=["-"], command=self.autocompletar_tecido)
            self.combo_modelo_novo.pack(fill="x", padx=20, pady=(0, 10))
            
            ctk.CTkLabel(self.frame_nova_cor, text="Tecido (Automático/Manual):", font=("Roboto", 12)).pack(anchor="w", padx=20)
            self.entry_tecido = ctk.CTkEntry(self.frame_nova_cor)
            self.entry_tecido.pack(fill="x", padx=20, pady=(0, 10))
            
            ctk.CTkLabel(self.frame_nova_cor, text="Nome Comercial da Cor:", font=("Roboto", 12)).pack(anchor="w", padx=20)
            self.entry_cor_comercial = ctk.CTkEntry(self.frame_nova_cor, placeholder_text="Ex: Azul Intense")
            self.entry_cor_comercial.pack(fill="x", padx=20, pady=(0, 10))
            
            ctk.CTkLabel(self.frame_nova_cor, text="Nome no Sistema (Parênteses):", font=("Roboto", 12)).pack(anchor="w", padx=20)
            self.entry_cor_sistema = ctk.CTkEntry(self.frame_nova_cor, placeholder_text="Ex: Azul Dinamarca")
            self.entry_cor_sistema.pack(fill="x", padx=20, pady=(0, 20))
            
            self.btn_salvar_cor = ctk.CTkButton(self.frame_nova_cor, text="💾 SALVAR COR NO EXCEL", height=45, font=("Roboto", 14, "bold"), fg_color="#2fa572", hover_color="#237a54", command=self.salvar_nova_cor)
            self.btn_salvar_cor.pack(fill="x", padx=20, pady=(0, 20))
            
            # AUTO-CARREGAMENTO
            self.after(500, self.carregar_planilha_automatica)

        # --- Funções do Gerenciador de Cores ---
        def carregar_planilha_automatica(self):
            if not os.path.exists(self.caminho_arquivo_cores):
                messagebox.showwarning("Aviso", f"Arquivo não encontrado no caminho automático:\n{self.caminho_arquivo_cores}\n\nPor favor, verifique se o caminho no código está correto, ou se o pendrive/rede Z: está conectado.", parent=self)
                return
            
            try:
                xl = pd.ExcelFile(self.caminho_arquivo_cores)
                self.dados_cores = {aba: pd.read_excel(xl, sheet_name=aba) for aba in xl.sheet_names}
                abas = list(self.dados_cores.keys())
                
                self.menu_abas.configure(values=abas)
                self.combo_modelo_novo.configure(values=abas)
                
                if abas:
                    self.menu_abas.set(abas[0])
                    self.combo_modelo_novo.set(abas[0])
                    self.exibir_cores_aba(abas[0])
                    self.autocompletar_tecido(abas[0])
                    
                self.lbl_arquivo.configure(text=f"Base Conectada e Sincronizada: {self.caminho_arquivo_cores}", text_color="#2fa572")
            except Exception as e:
                messagebox.showerror("Erro Crítico", f"Falha ao ler a base de cores automaticamente:\n{e}", parent=self)

        def extrair_nomes(self, cor_bruta):
            m = re.search(r'\((.*?)\)', cor_bruta)
            if m:
                sistema = m.group(1).strip()
                comercial = cor_bruta.replace(f"({m.group(1)})", "").strip()
                return comercial, sistema
            return cor_bruta, "-"
                
        def exibir_cores_aba(self, aba):
            if aba not in self.dados_cores: return
            df = self.dados_cores[aba]
            
            self.tabela.delete(*self.tabela.get_children())
            
            for _, row in df.iterrows():
                colunas = list(df.columns)
                modelo = str(row[colunas[0]]) if len(colunas) > 0 else aba
                tecido = str(row[colunas[1]]) if len(colunas) > 1 else ""
                cor_bruta = str(row[colunas[2]]) if len(colunas) > 2 else ""
                
                if pd.isna(cor_bruta) or cor_bruta.upper() == "NAN": continue
                
                cor_comercial, cor_sistema = self.extrair_nomes(cor_bruta)
                self.tabela.insert("", "end", values=(modelo, tecido, cor_comercial, f"({cor_sistema})"))
                
        def autocompletar_tecido(self, aba):
            if aba in self.dados_cores:
                df = self.dados_cores[aba]
                if not df.empty and len(df.columns) > 1:
                    tecido = str(df.iloc[0, 1])
                    if tecido.upper() != "NAN":
                        self.entry_tecido.delete(0, 'end')
                        self.entry_tecido.insert(0, tecido)
                        
        def criar_nova_aba(self):
            novo_modelo = self.entry_novo_modelo_aba.get().strip()
            
            if not novo_modelo:
                messagebox.showwarning("Atenção", "Digite um nome para o novo Modelo/Aba!", parent=self)
                return
                
            if novo_modelo in self.dados_cores:
                messagebox.showwarning("Atenção", f"Já existe uma aba chamada '{novo_modelo}'!", parent=self)
                return
                
            if not os.path.exists(self.caminho_arquivo_cores):
                messagebox.showerror("Erro", "Arquivo base não encontrado para edição.", parent=self)
                return
                
            try:
                wb = load_workbook(self.caminho_arquivo_cores)
                
                # Cria a nova aba no final
                ws = wb.create_sheet(title=novo_modelo)
                
                # Configurar Estilos do Cabeçalho (Igual à sua base)
                verde_fundo = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
                fonte_negrito = Font(bold=True)
                alinhamento = Alignment(horizontal="center", vertical="center")
                thin = Side(border_style="thin", color="000000")
                borda = Border(top=thin, left=thin, right=thin, bottom=thin)
                
                # Escrever cabeçalho padrão
                headers = ["Modelo", "Tecido (Nome no Sistema)", "Cor Disponível"]
                for col_idx, col_name in enumerate(headers, start=1):
                    celula = ws.cell(row=1, column=col_idx, value=col_name)
                    celula.fill = verde_fundo
                    celula.font = fonte_negrito
                    celula.alignment = alinhamento
                    celula.border = borda
                    
                # Ajustar largura das colunas
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 50
                
                wb.save(self.caminho_arquivo_cores)
                
                # Atualizar Sistema Virtualmente
                self.dados_cores[novo_modelo] = pd.DataFrame(columns=headers)
                abas = list(self.dados_cores.keys())
                self.menu_abas.configure(values=abas)
                self.combo_modelo_novo.configure(values=abas)
                
                # Selecionar a nova aba
                self.menu_abas.set(novo_modelo)
                self.combo_modelo_novo.set(novo_modelo)
                self.exibir_cores_aba(novo_modelo)
                
                self.entry_novo_modelo_aba.delete(0, 'end')
                
                messagebox.showinfo("Sucesso", f"O modelo/aba '{novo_modelo}' foi criado com sucesso no Excel!", parent=self)
                
            except PermissionError:
                messagebox.showerror("Erro de Permissão", "O Excel está ABERTO! Feche o arquivo para o sistema conseguir criar a aba.", parent=self)
            except Exception as e:
                messagebox.showerror("Erro Crítico", f"Erro ao tentar criar nova aba:\n{traceback.format_exc()}", parent=self)
                    
        def salvar_nova_cor(self):
            aba = self.combo_modelo_novo.get()
            tecido = self.entry_tecido.get().strip()
            comercial = self.entry_cor_comercial.get().strip()
            sistema = self.entry_cor_sistema.get().strip()
            
            if aba == "-" or not aba:
                messagebox.showwarning("Aviso", "Selecione um Modelo/Aba para salvar a cor!", parent=self)
                return
            
            if not comercial or not sistema:
                messagebox.showwarning("Aviso", "Preencha o nome Comercial e o nome no Sistema!", parent=self)
                return
                
            # Formatação obrigatória: Nome Comercial (Nome no Sistema)
            cor_final = f"{comercial} ({sistema})"
            
            try:
                wb = load_workbook(self.caminho_arquivo_cores)
                if aba not in wb.sheetnames:
                    messagebox.showerror("Erro", f"Aba '{aba}' não encontrada no arquivo.", parent=self)
                    return
                ws = wb[aba]
                
                nova_linha = ws.max_row + 1
                
                ws.cell(row=nova_linha, column=1, value=aba)
                ws.cell(row=nova_linha, column=2, value=tecido)
                ws.cell(row=nova_linha, column=3, value=cor_final)
                
                thin = Side(border_style="thin", color="000000")
                borda = Border(top=thin, left=thin, right=thin, bottom=thin)
                for col in range(1, 4):
                    ws.cell(row=nova_linha, column=col).border = borda
                
                wb.save(self.caminho_arquivo_cores)
                
                # Atualizar a tabela virtual na hora
                colunas = list(self.dados_cores[aba].columns)
                if len(colunas) >= 3:
                    nova_linha_df = {colunas[0]: aba, colunas[1]: tecido, colunas[2]: cor_final}
                    df_temp = pd.DataFrame([nova_linha_df])
                    self.dados_cores[aba] = pd.concat([self.dados_cores[aba], df_temp], ignore_index=True)
                
                if self.menu_abas.get() == aba:
                    self.exibir_cores_aba(aba)
                    
                self.entry_cor_comercial.delete(0, 'end')
                self.entry_cor_sistema.delete(0, 'end')
                
                messagebox.showinfo("Sucesso", f"Cor cadastrada!\n\nFoi adicionado: {cor_final}", parent=self)
                
            except PermissionError:
                messagebox.showerror("Erro de Permissão", "O Excel base das cores está ABERTO! Feche o arquivo para o sistema conseguir salvar.", parent=self)
            except Exception as e:
                messagebox.showerror("Erro Crítico", f"Erro ao tentar escrever no Excel:\n{traceback.format_exc()}", parent=self)


    # =======================================================
    # MÓDULO 1: SETOR DE PCP (CALCULADORA DE GRADE NA TELA)
    # =======================================================
    class JanelaPCP(ctk.CTkToplevel):
        def __init__(self, master):
            super().__init__(master)
            self.title("TEXAS FARM - Setor PCP")
            self.geometry("1300x950")
            self.protocol("WM_DELETE_WINDOW", self.fechar_e_voltar)

            self.caminho_arquivo = None
            self.abas_disponiveis = {}
            self.lotes_na_aba = {}
            self.lista_cores_atual = []
            self.fixado = False
            self.tamanhos_alvo = ['PP', 'P', 'M', 'G', 'GG', 'XG', 'G1', 'G2', 'G3', 'G4', '2', '4', '6', '8', '10', '12', '14']

            self.frame_topo = ctk.CTkFrame(self, fg_color="transparent")
            self.frame_topo.pack(pady=10, fill="x")
            
            self.btn_fixar = ctk.CTkButton(self.frame_topo, text="📌 PINAR JANELA", fg_color="#4a4a4a", command=self.alternar_fixar, width=150)
            self.btn_fixar.pack(side="left", padx=20)
            
            ctk.CTkLabel(self.frame_topo, text="📊 PAINEL DO PCP (CALCULADORA DE GRADE)", font=("Roboto", 20, "bold"), text_color="#3b8ed0").pack(side="left", expand=True)

            self.frame_main = ctk.CTkScrollableFrame(self)
            self.frame_main.pack(fill="both", expand=True, padx=20, pady=10)

            self.frame_excel = ctk.CTkFrame(self.frame_main)
            self.frame_excel.pack(pady=5, fill="x", padx=5)

            self.frame_botoes_acao = ctk.CTkFrame(self.frame_excel, fg_color="transparent")
            self.frame_botoes_acao.pack(pady=10)

            self.btn_excel = ctk.CTkButton(self.frame_botoes_acao, text="📂 1. LER PLANILHA DE PRODUÇÃO", fg_color="#1D6F42", hover_color="#144d2e", command=self.importar_excel)
            self.btn_excel.pack(side="left", padx=10)

            self.btn_cores = ctk.CTkButton(self.frame_botoes_acao, text="🎨 2. GERENCIAR CORES (BASE)", fg_color="#8B4513", hover_color="#5C2E0B", command=self.abrir_gerenciador_cores)
            self.btn_cores.pack(side="left", padx=10)

            self.frame_seletores = ctk.CTkFrame(self.frame_excel, fg_color="transparent")
            self.frame_seletores.pack(fill="x", padx=10, pady=5)

            self.menu_abas = ctk.CTkOptionMenu(self.frame_seletores, values=["-"], command=self.processar_aba_selecionada)
            self.menu_abas.grid(row=0, column=1, padx=5)

            self.entry_busca = ctk.CTkEntry(self.frame_seletores, placeholder_text="🔎 Buscar Lote...", width=150)
            self.entry_busca.grid(row=0, column=2, padx=5)
            self.entry_busca.bind("<KeyRelease>", self.filtrar_lotes)

            self.menu_lotes = ctk.CTkOptionMenu(self.frame_seletores, values=["-"], command=self.carregar_dados_lote, width=200)
            self.menu_lotes.grid(row=0, column=3, padx=5)

            self.frame_grades = ctk.CTkFrame(self.frame_main, fg_color="transparent")
            self.frame_grades.pack(pady=10)
            self.entradas_grade = {}
            self.setup_grade_ui()

            self.btn_calc = ctk.CTkButton(self.frame_main, text="3. CALCULAR GRADE EXATA", height=50, command=self.executar_grade_pcp, fg_color="#3b8ed0")
            self.btn_calc.pack(pady=20)

            self.frame_res = ctk.CTkFrame(self.frame_main)
            self.frame_res.pack(pady=5, fill="both", expand=True)

            self.label_soma_total = ctk.CTkLabel(self.frame_res, text="TOTAL DO LOTE: 0 PEÇAS", font=("Roboto", 18, "bold"))
            self.label_soma_total.pack(pady=10)

            style = ttk.Style()
            style.theme_use("clam")
            style.configure("Treeview", font=("Roboto", 12), rowheight=30, background="#2b2b2b", foreground="white")
            style.configure("Treeview.Heading", font=("Roboto", 13, "bold"), background="#3b8ed0", foreground="white")
            style.map("Treeview", background=[('selected', '#1f538d')])

            self.colunas = ("REF", "COR") + tuple(self.tamanhos_alvo) + ("TOTAL",)
            self.tabela = ttk.Treeview(self.frame_res, columns=self.colunas, show="headings")
            for col in self.colunas:
                self.tabela.heading(col, text=col)
                largura = 120 if col in ["REF", "COR"] else 55
                self.tabela.column(col, width=largura, anchor="center")
            self.tabela.pack(fill="both", expand=True)

        def abrir_gerenciador_cores(self):
            JanelaGerenciadorCores(self)

        def alternar_fixar(self): 
            self.fixado = not self.fixado
            self.attributes("-topmost", self.fixado)
            self.btn_fixar.configure(fg_color="#2fa572" if self.fixado else "#4a4a4a")

        def fechar_e_voltar(self): 
            self.master.deiconify()
            self.destroy()

        def setup_grade_ui(self):
            ctk.CTkLabel(self.frame_grades, text="ADULTO / PLUS", font=("Roboto", 12, "bold"), text_color="#3b8ed0").grid(row=0, column=0, columnspan=10, sticky="w", padx=5, pady=(5,2))
            for i, tam in enumerate(self.tamanhos_alvo[:10]):
                ctk.CTkLabel(self.frame_grades, text=tam, font=("Roboto", 12)).grid(row=1, column=i, padx=6)
                self.entradas_grade[tam] = ctk.CTkEntry(self.frame_grades, width=55, justify="center")
                self.entradas_grade[tam].insert(0, "0")
                self.entradas_grade[tam].grid(row=2, column=i, padx=2, pady=2)

            ctk.CTkLabel(self.frame_grades, text="INFANTIL (2-14)", font=("Roboto", 12, "bold"), text_color="#3b8ed0").grid(row=3, column=0, columnspan=10, sticky="w", padx=5, pady=(10,2))
            for i, tam in enumerate(self.tamanhos_alvo[10:]):
                ctk.CTkLabel(self.frame_grades, text=tam, font=("Roboto", 12)).grid(row=4, column=i, padx=6)
                self.entradas_grade[tam] = ctk.CTkEntry(self.frame_grades, width=55, justify="center")
                self.entradas_grade[tam].insert(0, "0")
                self.entradas_grade[tam].grid(row=5, column=i, padx=2, pady=2)

        def importar_excel(self):
            c = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
            if not c: return
            self.caminho_arquivo = c
            try:
                self.abas_disponiveis = pd.read_excel(c, sheet_name=None, header=None)
                n = list(self.abas_disponiveis.keys())
                self.menu_abas.configure(values=n if n else ["-"])
                if n:
                    self.menu_abas.set(n[0])
                    self.processar_lotes_da_planilha(n[0])
                messagebox.showinfo("Sucesso", "Planilha carregada com sucesso!")
            except Exception as e: 
                messagebox.showerror("Erro", f"Erro ao ler arquivo: {e}")

        def processar_aba_selecionada(self, n_aba):
            self.processar_lotes_da_planilha(n_aba)
            
        def processar_lotes_da_planilha(self, n_aba):
            df = self.abas_disponiveis[n_aba]
            self.lotes_na_aba = {}
            lote_atual = None
            dados_l = []

            for i, linha in df.iterrows():
                l_str = " ".join([str(val) for val in linha.values if pd.notna(val)]).upper()
                
                if "LOTE:" in l_str:
                    lote_atual = str(linha[0]).strip().upper() if "LOTE:" in str(linha[0]).upper() else str(linha[1]).strip().upper()
                    grade_lote = {t: 0 for t in self.tamanhos_alvo}
                    
                    for r in range(i, min(i + 20, len(df))):
                        for col_idx in range(2, min(7, len(df.columns))):
                            celula = str(df.iloc[r, col_idx]).upper()
                            for t in self.tamanhos_alvo:
                                m = re.search(rf"\b{t}\b\s*[=:\-]\s*(\d+)", celula)
                                if m and grade_lote[t] == 0: 
                                    grade_lote[t] = int(m.group(1))

                    self.lotes_na_aba[lote_atual] = {"itens": [], "aba": n_aba, "grade": grade_lote}
                    dados_l = self.lotes_na_aba[lote_atual]["itens"]
                
                elif lote_atual and not pd.isna(linha[1]) and not pd.isna(linha[2]):
                    cor = str(linha[1]).upper()
                    if "COR" in cor or "TOTAL" in cor or "QUANT" in str(linha[2]).upper(): continue
                    try:
                        try:
                            r_val = str(linha[0]).replace(',', '.').strip()
                            rolos_qte = int(round(float(r_val))) if r_val.upper() not in ["NAN", "NONE", ""] else 0
                        except:
                            rolos_qte = 0
                            
                        dados_l.append({
                            'rolos': rolos_qte,
                            'cor': str(linha[1]).strip(),
                            'ref': str(linha[3]).strip() if pd.notna(linha[3]) else "Sem Ref",
                            'total': int(round(float(str(linha[2]).replace(',','.'))))
                        })
                    except: continue

            n = list(self.lotes_na_aba.keys())
            self.menu_lotes.configure(values=n if n else ["-"])
            if n:
                self.menu_lotes.set(n[0])
                self.carregar_dados_lote(n[0])
            else:
                self.menu_lotes.set("-")
                
        def carregar_dados_lote(self, n):
            if n == "-": return
            lote = self.lotes_na_aba.get(n)
            if not lote: return
            self.lista_cores_atual = lote["itens"]
            
            for tam in self.tamanhos_alvo:
                valor = lote.get("grade", {}).get(tam, 0)
                self.entradas_grade[tam].delete(0, 'end')
                self.entradas_grade[tam].insert(0, str(valor))
                
            self.tabela.delete(*self.tabela.get_children())
            soma = sum([i['total'] for i in self.lista_cores_atual])
            self.label_soma_total.configure(text=f"TOTAL DO LOTE: {soma} PEÇAS")

        def filtrar_lotes(self, e):
            t = self.entry_busca.get().upper()
            f = [l for l in self.lotes_na_aba.keys() if t in l]
            self.menu_lotes.configure(values=f if f else ["-"])
            if f: 
                self.menu_lotes.set(f[0])
                self.carregar_dados_lote(f[0])
            else:
                self.menu_lotes.set("-")

        def executar_grade_pcp(self):
            if not self.lista_cores_atual: return
            try:
                pesos = {k: int(v.get().strip() or 0) for k, v in self.entradas_grade.items()}
                soma_pesos = sum(pesos.values())
                if soma_pesos == 0: return

                self.tabela.delete(*self.tabela.get_children())
                soma_geral = 0

                for item in self.lista_cores_atual:
                    total_cor = item['total']
                    refs = [r.strip() for r in str(item['ref']).split('/') if r.strip()]
                    refs = refs if refs else ["Sem Ref"]
                    
                    grade_cor = {t: 0 for t in self.tamanhos_alvo}
                    fracoes = {t: 0.0 for t in self.tamanhos_alvo}
                    
                    for t, p in pesos.items():
                        if p > 0:
                            valor_exato = (total_cor * p) / soma_pesos
                            grade_cor[t] = math.floor(valor_exato)
                            fracoes[t] = valor_exato - math.floor(valor_exato)
                    
                    sobra = total_cor - sum(grade_cor.values())
                    if sobra > 0:
                        tams_ativos = [t for t, p in pesos.items() if p > 0]
                        if tams_ativos:
                            tams_ord = sorted(tams_ativos, key=lambda x: fracoes[x], reverse=True)
                            for i in range(sobra): 
                                grade_cor[tams_ord[i % len(tams_ord)]] += 1
                    
                    for idx, ref in enumerate(refs):
                        div = len(refs)
                        tot_ref = {t: (v // div) + (1 if idx < (v % div) else 0) for t, v in grade_cor.items()}
                        soma_ref_final = sum(tot_ref.values())
                        
                        self.tabela.insert("", "end", values=[ref, item['cor']] + [tot_ref[t] for t in self.tamanhos_alvo] + [soma_ref_final])
                        soma_geral += soma_ref_final
                
                self.label_soma_total.configure(text=f"TOTAL CALCULADO: {soma_geral} PEÇAS")
            except Exception as e: 
                messagebox.showerror("Erro de Cálculo", traceback.format_exc())

    # =======================================================
    # MÓDULO 2: SETOR ALMOXARIFADO (ESTOQUE)
    # =======================================================
    class JanelaEstoque(ctk.CTkToplevel):
        def __init__(self, master):
            super().__init__(master)
            self.title("TEXAS FARM - Setor Almoxarifado/Estoque")
            self.geometry("1100x800")
            self.protocol("WM_DELETE_WINDOW", self.fechar_e_voltar)
            
            self.arquivo_memoria = r"Z:\TXF\Software_PCP_TexasFarm\historico_lotes_mt.txt"
            self.caminho_arquivo = None
            self.lotes_processados_agora = set()
            self.novo_estoque_calculado = {}

            # --- CONSTRUÇÃO DA INTERFACE ---
            self.frame_topo = ctk.CTkFrame(self, fg_color="transparent")
            self.frame_topo.pack(pady=10, fill="x")
            ctk.CTkLabel(self.frame_topo, text="📦 PAINEL DO ALMOXARIFADO", font=("Roboto", 20, "bold"), text_color="#d08d3b").pack()

            self.frame_controles = ctk.CTkFrame(self)
            self.frame_controles.pack(pady=5, padx=20, fill="x")
            ctk.CTkButton(self.frame_controles, text="📂 1. ABRIR PLANILHA", command=self.importar_excel, fg_color="#1D6F42", font=("Roboto", 13, "bold")).pack(side="left", padx=5, pady=10)
            ctk.CTkButton(self.frame_controles, text="🔄 2. ESTOQUE ATUAL", command=self.ver_estoque, fg_color="#2fa572", font=("Roboto", 13, "bold")).pack(side="left", padx=5, pady=10)
            ctk.CTkButton(self.frame_controles, text="🧵 3. ATUALIZAR ESTOQUE", command=self.calcular_saldo, fg_color="#1f538d", font=("Roboto", 13, "bold")).pack(side="left", padx=5, pady=10)
            self.btn_confirmar = ctk.CTkButton(self.frame_controles, text="✔️ 4. CONFIRMAR E SALVAR NO EXCEL", command=self.salvar_historico_e_excel, fg_color="#8b0000", state="disabled", font=("Roboto", 13, "bold"))
            self.btn_confirmar.pack(side="left", padx=5, pady=10)

            self.frame_seletores_est = ctk.CTkFrame(self)
            self.frame_seletores_est.pack(pady=5, padx=20, fill="x")
            
            ctk.CTkLabel(self.frame_seletores_est, text="Aba do Lote:", font=("Roboto", 12)).grid(row=0, column=0, padx=5, pady=5)
            self.menu_aba_lote = ctk.CTkOptionMenu(self.frame_seletores_est, values=["-"])
            self.menu_aba_lote.grid(row=0, column=1, padx=5, pady=5)
            
            ctk.CTkLabel(self.frame_seletores_est, text="Aba do Estoque:", font=("Roboto", 12)).grid(row=0, column=2, padx=5, pady=5)
            self.menu_aba_est = ctk.CTkOptionMenu(self.frame_seletores_est, values=["-"])
            self.menu_aba_est.grid(row=0, column=3, padx=5, pady=5)

            self.lbl_status = ctk.CTkLabel(self, text="Selecione a planilha...", text_color="gray", font=("Roboto", 14))
            self.lbl_status.pack(pady=5)

            self.frame_tabela = ctk.CTkFrame(self)
            self.frame_tabela.pack(fill="both", expand=True, padx=20, pady=10)

            style = ttk.Style()
            style.theme_use("clam")
            style.configure("Treeview", font=("Roboto", 13), rowheight=30, background="#2b2b2b", foreground="white")
            style.configure("Treeview.Heading", font=("Roboto", 14, "bold"), background="#d08d3b", foreground="white")

            self.tabela = ttk.Treeview(self.frame_tabela, show="headings")
            self.tabela.pack(fill="both", expand=True)

        def fechar_e_voltar(self): 
            self.master.deiconify()
            self.destroy()

        def importar_excel(self):
            c = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
            self.caminho_arquivo = c
            if c: 
                self.lbl_status.configure(text=f"Carregado: {c.split('/')[-1]}", text_color="white")
                try:
                    xl = pd.ExcelFile(c)
                    nomes_abas = xl.sheet_names
                    self.menu_aba_lote.configure(values=nomes_abas if nomes_abas else ["-"])
                    self.menu_aba_est.configure(values=nomes_abas if nomes_abas else ["-"])
                    
                    sel_lote = next((s for s in nomes_abas if any(x in s.upper() for x in ["LOTE", "PROGRAMA", "PRODU", "CORTES"])), nomes_abas[0])
                    sel_est = next((s for s in nomes_abas if any(x in s.upper() for x in ["REFER", "ROLO", "ESTOQUE", "CONTROLE", "MOLETOM", "MALHA"])), nomes_abas[0])
                    
                    self.menu_aba_lote.set(sel_lote)
                    self.menu_aba_est.set(sel_est)
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao ler as abas do arquivo: {e}")

        def extrair_estoque(self):
            if not self.caminho_arquivo: return None
            aba_est = self.menu_aba_est.get()
            if aba_est == "-": return None
            try:
                df = pd.read_excel(self.caminho_arquivo, sheet_name=aba_est, header=None)
                estoque = {}
                cc, cq = -1, -1
                
                for r_idx, row in df.iterrows():
                    for c_idx, val in enumerate(row):
                        v = str(val).strip().upper()
                        if v == "COR" or v == "CORES": cc = c_idx
                        if "QTE" in v or "ROLO" in v or "QTD" in v or "QUANT" in v: cq = c_idx
                    if cc != -1 and cq != -1: break
                    
                for _, row in df.iterrows():
                    cor_bruta = str(row[cc]).upper()
                    cor_limpa = re.sub(r'\(.*?\)', '', cor_bruta).strip()
                    if cor_limpa in ["COR", "NAN", "NONE", "", "TOTAL"] or pd.isna(row[cc]): continue
                    try:
                        v = str(row[cq]).replace(',', '.').strip()
                        if v.upper() not in ["NAN", "NONE", ""]: estoque[cor_limpa] = float(v)
                    except: pass
                return estoque
            except PermissionError:
                messagebox.showerror("Erro", "O Excel está aberto! Feche o arquivo para prosseguir.")
                return None
            except Exception as e: 
                messagebox.showerror("Erro", f"Falha ao ler estoque: {e}")
                return None

        def ver_estoque(self):
            est = self.extrair_estoque()
            if not est: return
            
            self.tabela.delete(*self.tabela.get_children())
            self.tabela["columns"] = ("COR", "QTD SALVA NO EXCEL")
            for c in self.tabela["columns"]: 
                self.tabela.heading(c, text=c)
                self.tabela.column(c, anchor="center")
                
            total = 0
            for cor, q in est.items(): 
                self.tabela.insert("", "end", values=(cor, q))
                total += q
            self.tabela.insert("", "end", values=("TOTAL GERAL", total))
            self.lbl_status.configure(text="Estoque atualizado em tempo real da planilha!", text_color="#2fa572")

        def calcular_saldo(self):
            est = self.extrair_estoque()
            if not est: return
            
            aba_lote = self.menu_aba_lote.get()
            if aba_lote == "-": return
                
            df_lote = pd.read_excel(self.caminho_arquivo, sheet_name=aba_lote, header=None)
            cons, ign = {}, set()
            
            try:
                os.makedirs(os.path.dirname(self.arquivo_memoria), exist_ok=True)
                if os.path.exists(self.arquivo_memoria):
                    with open(self.arquivo_memoria, "r", encoding="utf-8") as f: ign = set(f.read().splitlines())
            except: pass
            
            l_at, self.lotes_processados_agora, cores_erro = None, set(), []
            for _, linha in df_lote.iterrows():
                l_str = " ".join([str(val).upper() for val in linha.values if pd.notna(val)]).upper()
                if "LOTE:" in l_str:
                    for cel in linha.values:
                        if pd.notna(cel) and "LOTE:" in str(cel).upper(): 
                            l_at = str(cel).strip().upper()
                            break
                    continue
                if "TOTAL" in l_str: l_at = None; continue
                
                if l_at and l_at.startswith("MT") and l_at not in ign:
                    cor_bruta = str(linha[1]).upper()
                    cor_limpa = re.sub(r'\(.*?\)', '', cor_bruta).strip()
                    if cor_limpa in ["COR", "NAN", "NONE", "", "TOTAL"] or pd.isna(linha[1]): continue
                    if cor_limpa not in est:
                        if cor_limpa not in cores_erro: cores_erro.append(cor_limpa)
                        continue

                    self.lotes_processados_agora.add(l_at)
                    try:
                        r = float(str(linha[0]).replace(',', '.').strip())
                        if not pd.isna(r) and r > 0: cons[cor_limpa] = cons.get(cor_limpa, 0.0) + r
                    except: pass

            if cores_erro: messagebox.showwarning("Cores Ausentes", f"Ignoradas por não estarem no Estoque:\n\n" + "\n".join(cores_erro))

            self.tabela.delete(*self.tabela.get_children())
            self.tabela["columns"] = ("COR", "INI", "CONS", "NOVO SALDO")
            for c in self.tabela["columns"]: 
                self.tabela.heading(c, text=c)
                self.tabela.column(c, anchor="center")
                
            d_t = []
            self.novo_estoque_calculado = {}
            for cor in sorted(set(list(est.keys()) + list(cons.keys()))):
                q_i, q_c = est.get(cor, 0.0), cons.get(cor, 0.0)
                if q_i > 0 or q_c > 0:
                    novo_saldo = q_i - q_c
                    self.tabela.insert("", "end", values=(cor, q_i, q_c, novo_saldo))
                    d_t.append(novo_saldo)
                    self.novo_estoque_calculado[cor] = novo_saldo
            
            if self.novo_estoque_calculado:
                self.tabela.insert("", "end", values=("TOTAL GERAL", "", "", sum(d_t)))
                self.btn_confirmar.configure(state="normal")
                self.lbl_status.configure(text="Saldos prontos! Clique em CONFIRMAR para gravar no Excel.", text_color="#d08d3b")

        def salvar_historico_e_excel(self):
            if not self.lotes_processados_agora or not self.novo_estoque_calculado: return
            if messagebox.askyesno("Confirmar e Salvar", "Deseja gravar o novo estoque na planilha e salvar o histórico TXT?"):
                try:
                    wb = load_workbook(self.caminho_arquivo)
                    aba_est = self.menu_aba_est.get()
                    if aba_est not in wb.sheetnames: raise Exception("Aba de Estoque selecionada inválida.")
                    
                    ws = wb[aba_est]
                    cc, cq, header_row = -1, -1, 1
                    
                    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                        for c_idx, val in enumerate(row):
                            v = str(val).strip().upper() if val else ""
                            if v == "COR" or v == "CORES": cc = c_idx
                            if "QTE" in v or "ROLO" in v or "QTD" in v or "QUANT" in v: cq = c_idx
                        if cc != -1 and cq != -1: header_row = r_idx + 1; break
                            
                    if cc == -1 or cq == -1: raise Exception("Colunas COR e QTD não identificadas.")

                    soma_total_excel = 0
                    linha_total = None

                    for row in ws.iter_rows(min_row=header_row+1):
                        cor_bruta = str(row[cc].value).upper() if row[cc].value else ""
                        cor_limpa = re.sub(r'\(.*?\)', '', cor_bruta).strip()

                        if cor_limpa in self.novo_estoque_calculado: row[cq].value = self.novo_estoque_calculado[cor_limpa]
                        if "TOTAL" in cor_limpa: linha_total = row
                        elif row[cq].value is not None and str(row[cq].value).replace('.','',1).isdigit(): soma_total_excel += float(row[cq].value)

                    if linha_total: linha_total[cq].value = soma_total_excel

                    wb.save(self.caminho_arquivo)

                    caminho_dir = os.path.dirname(self.arquivo_memoria)
                    if not os.path.exists(caminho_dir):
                        try: os.makedirs(caminho_dir, exist_ok=True)
                        except Exception: self.arquivo_memoria = "historico_lotes_mt_local.txt"

                    with open(self.arquivo_memoria, "a", encoding="utf-8") as f:
                        for l in self.lotes_processados_agora: f.write(l + "\n")
                    
                    self.btn_confirmar.configure(state="disabled")
                    self.lotes_processados_agora.clear()
                    self.novo_estoque_calculado.clear()
                    
                    self.lbl_status.configure(text="Estoque, Total e Histórico salvos com sucesso!", text_color="#1D6F42")
                    messagebox.showinfo("Sucesso", "Estoque gravado no Excel com TOTAL atualizado!")
                    self.ver_estoque()

                except PermissionError: messagebox.showerror("Erro Crítico", "O Excel está ABERTO! Você precisa fechar o arquivo antes de tentar gravar.")
                except Exception as e: messagebox.showerror("Erro de Gravação", traceback.format_exc())

    # =======================================================
    # LAUNCHER HUB
    # =======================================================
    class LauncherTexasFarm(ctk.CTk):
        def __init__(self):
            super().__init__()
            self.title("TEXAS FARM - ERP Hub")
            sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
            self.geometry(f"500x400+{(sw-500)//2}+{(sh-400)//2}")
            ctk.CTkLabel(self, text="TEXAS FARM", font=("Roboto", 28, "bold")).pack(pady=(40, 5))
            ctk.CTkButton(self, text="⚙️ PCP", height=70, width=300, font=("Roboto", 15, "bold"), command=self.abrir_pcp).pack(pady=10)
            ctk.CTkButton(self, text="📦 ESTOQUE DE ROLOS", height=70, width=300, font=("Roboto", 15, "bold"), fg_color="#d08d3b", command=self.abrir_estoque).pack(pady=10)
        
        def abrir_pcp(self): self.withdraw(); JanelaPCP(self)
        def abrir_estoque(self): self.withdraw(); JanelaEstoque(self)

    if __name__ == "__main__": 
        LauncherTexasFarm().mainloop()

except Exception as e:
    root = tk.Tk()
    root.withdraw()
    tkmb.showerror("Erro Fatal", f"Ocorreu um erro ao iniciar o sistema.\nCertifique-se de que as bibliotecas estão instaladas.\n\nDetalhes do erro:\n{traceback.format_exc()}")
